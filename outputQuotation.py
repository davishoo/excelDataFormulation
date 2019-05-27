import openpyxl, pprint, logging, time, sys
from openpyxl.styles import Font
from datetime import datetime

logging.disable(logging.CRITICAL)
logging.debug('Start of program')

try:
    from hotelQuotation import quotationData
    print('Import database module: [hotelQuotation] ...')
except  ModuleNotFoundError:
    print('warning: no hotelQuotation module found!')
    sys.exit()


try:
    from hotelDatabase import hotelDB
    print('Import database module: [hotelDatabase] ...')
except:
    print('warning: no hotelDatabase module found!')
    sys.exit()


wb = openpyxl.Workbook()
sheet = wb['Sheet']


#fill out the head of the sheet
rowNum = 2

for project, v1 in quotationData.keys()    
    for quotationSource, v2 in quotationData[
        sheet.cell(row = rowNum, column = 1).value = quotationSource
        for hotel in quotation[project][quotationSource]            
            sheet.cell(row = rowNum, column = 1).value = hotelDB[hotel][-1]['hotelLocation']
            sheet.cell(row = rowNum, column = 2).value = hotel
            for roomType in quotation[project][quotationSource][hotel]
                sheet.cell(row = rowNum, column = 4).value = roomType
                for quoteDate in quotation[project][quotationSource][hotel][roomType]
                    sheet.cell(row = rowNum, column = 5).value = quoteDate


sheet['A1'] = 'hello'
wb.save('hotel_quotation.xlsx')






        
